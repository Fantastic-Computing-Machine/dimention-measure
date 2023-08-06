from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill,Border,Side
from openpyxl.drawing.image import Image

all_side_thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

all_side_thick_border = Border(left=Side(style='thick'),
                        right=Side(style='thick'),
                        top=Side(style='thick'),
                        bottom=Side(style='thick'))

top_thin_border = Border(top=Side(style='thin'))
left_thin_border = Border(left=Side(style='thin'))
right_thin_border = Border(right=Side(style='thin'))
bottom_thin_border = Border(bottom=Side(style='thin'))

top_right_thin_border = Border(top=Side(style='thin'), right=Side(style='thin'))
top_left_thin_border = Border(top=Side(style='thin'), left=Side(style='thin'))
bottom_right_thin_border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
bottom_left_thin_border = Border(bottom=Side(style='thin'), left=Side(style='thin'))

top_bottom_thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
top_bottom_right_thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
top_bottom_left_thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

def generate_Invoice():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Invoice"
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.sheet_properties.pageSetUpPr.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToWidth = 1
    sheet.sheet_properties.tabColor = "1072BA"
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.sheet_properties.outlinePr.summaryRight = False
    sheet.sheet_properties.filterMode = False
    sheet.sheet_properties.published = False
    sheet.sheet_properties.codeName = "Invoice"

    sheet.column_dimensions["A"].width = 35.67
    sheet.column_dimensions["B"].width = 9
    sheet.column_dimensions["C"].width = 9
    sheet.column_dimensions["D"].width = 9
    sheet.column_dimensions["E"].width = 13.33
    sheet.column_dimensions["F"].width = 3.33
    sheet.column_dimensions["G"].width = 16.78

    # Heading
    sheet.merge_cells('A1:G1')
    sheet["A1"].value = "TAX INVOICE"
    sheet.merge_cells("A4:A6")
    sheet["A7"].value = "Email: " 
    sheet["A8"].value = "Mobile: "
    sheet["E3"], sheet["F3"] = "Invoice No.", " : "
    sheet["E4"], sheet["F4"] = "Invoice Date", " : "
    sheet["E5"], sheet["F5"]= "GSTIN/UIN", " : "
    sheet["E6"], sheet["F6"] = "PAN No.", " : "
    sheet["E7"], sheet["F7"]= "Reference No.", " : "

    sheet.merge_cells('A9:D9')
    sheet["A9"].value = "Customer Name (Bill To)"

    sheet.merge_cells('A10:D10')
    sheet.merge_cells('A11:D12')
    sheet.merge_cells('E9:G9')
    sheet["E9"].value = "Project Address"

    sheet.merge_cells('E10:G12')

    sheet.merge_cells('A13:D13')
    sheet.merge_cells('E13:G13')
    sheet.merge_cells('A14:D14')
    sheet.merge_cells('E14:G14')
    sheet["A13"].value = "PAN No."

    sheet.append(["Description of Services", "HSN/SAC", "GST Rate", "Quantity","Rate","Per","Amount" ])

    sheet["A37"].value = "CGST"
    sheet["A38"].value = "SGST"
    sheet["F37"].value = "%"
    sheet["F38"].value = "%"

    sheet["E39"].value = "Total"

    sheet["E40"].value = "Total Tax"
    sheet["E41"].value = "Total Invoice Value"
    sheet["E42"].value = "Round off"

    sheet.merge_cells('A43:G43')
    sheet["A43"].value = "Amount Chargeable (in words) : "

    sheet["A44"].value = "Bank Details"

    sheet["A45"].value = "Name : "
    sheet["A46"].value = "Account No. : "
    sheet["A47"].value = "Bank Name : "
    sheet["A48"].value = "Branch : "
    sheet["A49"].value = "IFSC Code : "

    sheet["D44"].value = "For "
    sheet["D48"].value = "Authorized Signatory"

    sheet["D49"].value = "Note : The Bill has to be paid within 7 days from the date of invoice."

    # Borders
    sheet["A1"].border = sheet["B1"].border = sheet["C1"].border = sheet["D1"].border = sheet["E1"].border = sheet["F1"].border = sheet["G1"].border = bottom_thin_border
  
    for _ in range(1,50):
        sheet["H{}".format(_)].border = left_thin_border
    
    sheet["A50"].border = sheet["B50"].border = sheet["C50"].border = sheet["D50"].border = sheet["E50"].border = sheet["F50"].border = sheet["G50"].border = top_thin_border

    for _ in range(2,9):
        sheet["B{}".format(_)].border = sheet["E{}".format(_)].border = left_thin_border
    
    sheet["A9"].border = sheet["B9"].border = sheet["C9"].border  = sheet["E9"].border = sheet["F9"].border = sheet["G9"].border = top_bottom_thin_border
    sheet["D9"].border = top_bottom_right_thin_border

    sheet["A10"].border = sheet["B10"].border = sheet["C10"].border = sheet["D10"].border =  bottom_thin_border
    for _ in range(10,14):
        sheet["E{}".format(_)].border = left_thin_border

    sheet["A14"].border = sheet["B14"].border = sheet["C14"].border = sheet["D14"].border = sheet["F14"].border = sheet["G14"].border = top_thin_border
    sheet["E14"].border = top_left_thin_border

    sheet["A15"].border = sheet["B15"].border = sheet["C15"].border = sheet["D15"].border = sheet["F15"].border =sheet["E15"].border = sheet["G15"].border = all_side_thin_border

    for _ in range(16,39):
        sheet["A{}".format(_)].border = sheet["B{}".format(_)].border = sheet["C{}".format(_)].border = sheet["D{}".format(_)].border = sheet["E{}".format(_)].border = sheet["F{}".format(_)].border  = right_thin_border

    sheet["A39"].border = sheet["B39"].border = sheet["C39"].border = sheet["D39"].border = sheet["E39"].border = sheet["F39"].border = sheet["G39"].border = top_bottom_thin_border

    sheet["A43"].border = sheet["B43"].border = sheet["C43"].border = sheet["D43"].border = sheet["E43"].border = sheet["F43"].border = sheet["G43"].border = top_bottom_thin_border
    for _ in range(44,50):
        sheet["C{}".format(_)].border = right_thin_border
    
    # Color
    sheet["A1"].fill = PatternFill("solid", fgColor="D3D3D3")
    sheet["A15"].fill = sheet["B15"].fill = sheet["C15"].fill = sheet["D15"].fill = sheet["E15"].fill = sheet["F15"].fill = sheet["G15"].fill = PatternFill("solid", fgColor="D3D3D3")
    sheet["A39"].fill = sheet["B39"].fill = sheet["C39"].fill = sheet["D39"].fill = sheet["E39"].fill = sheet["F39"].fill = sheet["G39"].fill = PatternFill("solid", fgColor="D3D3D3")

    # Alignment
    sheet["A3"].alignment = Alignment(horizontal="left", wrap_text=True)
    sheet["A4"].alignment = Alignment(horizontal="left", wrap_text=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A10"].alignment = Alignment(horizontal="left", wrap_text=True)
    sheet["A15"].alignment = sheet["B15"].alignment = sheet["C15"].alignment = sheet["D15"].alignment = sheet["E15"].alignment = sheet["F15"].alignment = sheet["G15"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    for _ in range(16,37):
        sheet["A{}".format(_)].alignment = sheet["B{}".format(_)].alignment = sheet["C{}".format(_)].alignment = sheet["D{}".format(_)].alignment = sheet["E{}".format(_)].alignment = sheet["F{}".format(_)].alignment = Alignment(horizontal="left", wrap_text=True)

    sheet["A43"].alignment = Alignment(horizontal="left", wrap_text=True)
    sheet["E41"].alignment = Alignment(horizontal="right")
    sheet["A37"].alignment = sheet["A38"].alignment =  Alignment(horizontal="right")

    # Font
    
    sheet["A1"].font =  sheet["A9"].font = sheet["A10"].font = sheet["A43"].font = sheet["A44"].font = Font(size = 10,bold=True)
    sheet["A3"].font = Font(size = 12, bold=True)
    sheet["E3"].font = sheet["E4"].font = sheet["E5"].font = sheet["E6"].font = sheet["E7"].font = sheet["E9"].font = Font(size = 10,bold=True)

    sheet["A15"].font = sheet["B15"].font = sheet["C15"].font = sheet["D15"].font = sheet["E15"].font = sheet["F15"].font = sheet["G15"].font = Font(size = 10,bold=True)

    sheet["E39"].font = Font(size = 10,bold=True)
    sheet["A43"].font = sheet["A44"].font = sheet["D44"].font = sheet["E41"].font = sheet["E42"].font = Font(size = 10,bold=True)
    sheet["G40"].font = sheet["G41"].font = sheet["G42"].font = Font(size = 10,bold=True)
    sheet["E44"].font = Font(size = 10,bold=True)
    sheet["D49"].font = Font(size=7)

    return wb